from rest_framework import viewsets, filters
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from ..serializers.common_serializers import *
from ..models.columns import Column
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import action
from system import utils
import openpyxl
from django.db.models import Q
from system.models.dataset import DataTable, Data

# Generate Token Manually
def get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)
    
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
    }

#**********************************Function To Extract Data From Excel File******************************************#
def extracting_data(file):
    xl_data = []
    heading = []
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    max_col = sheet.max_column
    for i in range(1, max_col+1):
        cell_obj = sheet.cell(row = 1, column = i)
        header = utils.encode_api_name(cell_obj.value)
        heading.append(header)
    for row in sheet.iter_rows(min_row=2):
        data_dict={}
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        for i in range(len(row_data)):
            data_dict[heading[i]]=row_data[i]
        xl_data.append(data_dict)
    return(xl_data)

class ButtonViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Buttons to be modified.
    """
    queryset = Button.objects.all()
    serializer_class = ButtonSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
class CurrencyViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Currency to be modified.
    """
    queryset = Currency.objects.all()
    serializer_class = CurrencySerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    try:
                        serializer=CurrencySerializer(data=data_dict[i], context={'request':request})
                        if serializer.is_valid(raise_exception=True):
                            serializer.save()
                            count += 1
                    except Exception as e:
                        print("CURRENCY ERROR >>>> ",str(e)) 
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))
           
class TagViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Tag to be modified.
    """
    queryset = Tag.objects.all()
    serializer_class = TagSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
class LanguageViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Language to be modified.
    """
    queryset = Language.objects.all()
    serializer_class = LanguageSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

class CountryViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Country to be modified.
    """
    queryset = Country.objects.all()
    serializer_class = CountrySerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = {'system_name': ['exact', 'icontains'],'country_code': ['exact'], 'telephone_code':['exact'], 'currency':['exact']}
    ordering_fields = ("__all__")

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            symbol_selector = Selectors.objects.get(system_name__contains='country_symbol_position')
            money_selector = Selectors.objects.get(system_name__contains='country_money_format')
            date_selector = Selectors.objects.get(system_name__contains='country_date_format')
            time_selector = Selectors.objects.get(system_name__contains='country_time_format')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    currency=data_dict[i].get('currency')
                    country=data_dict[i].get('country')
                    currency_id = data_dict[i].pop('currency_id')
                    symbol_position = data_dict[i].pop('currency_symbol_position')
                    money_format = data_dict[i].pop('money_format')
                    data_format = data_dict[i].pop('date_format')
                    time_format = data_dict[i].pop('time_format')
                    if currency and country:
                        try:
                            currency_rec = Currency.objects.get(code=currency, id=currency_id)
                            symbol_rec = Choice.objects.filter(system_name=symbol_position, selector=symbol_selector.id)
                            money_rec = Choice.objects.filter(system_name=money_format, selector=money_selector.id)
                            date_rec = Choice.objects.filter(system_name=data_format, selector=date_selector.id)
                            time_rec = Choice.objects.filter(system_name=time_format, selector=time_selector.id)

                            data_dict[i]['currency']=currency_rec.id
                            if symbol_rec:
                                data_dict[i]['symbol_position']=symbol_rec.values()[0]['id']
                            if money_rec:
                                data_dict[i]['money_format']=money_rec.values()[0]['id']
                            if date_rec:
                                data_dict[i]['date_format']=date_rec.values()[0]['id']
                            if time_rec:
                                data_dict[i]['time_format']=time_rec.values()[0]['id']

                            serializer=CountrySerializer(data=data_dict[i], context={'request':request})
                            if serializer.is_valid(raise_exception=True):
                                serializer.save()
                                count += 1
                        except Exception as e:
                            print("COUNTRY ERROR >>>> ", str(e), data_dict[i])
                            pass
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))
        
class StateViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows State to be modified.
    """
    queryset = State.objects.all()
    serializer_class = StateSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                defective_data=[]
                for i in range(len(data_dict)):
                    country=data_dict[i].get('country')
                    state_name = data_dict[i].get('name')
                    if country and state_name:
                        try:
                            country_rec = Country.objects.get(Q(country__iname=country) | Q(country=country))
                            data_dict[i]['country']=country_rec.id
                            
                            state_check=State.objects.filter(system_name=state_name, country=country_rec.id)
                            if not state_check:
                                serializer=StateSerializer(data=data_dict[i], context={'request':request})
                                if serializer.is_valid(raise_exception=True):
                                    serializer.save()
                                    count += 1
                        except Exception as e:
                            defective_data.append(country)
                            print(str(e), data_dict[i])
                            pass
                if defective_data:
                    defective_data = {
                        "inavlid_country" : f"These {set(defective_data)} are the invalid country names."
                    }
                    return Response(utils.success_def(count,defective_data))
                else:
                    return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))
    
class StageViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Stage to be modified.
    """
    queryset = Stage.objects.all()
    serializer_class = StageSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                defective_data=[]
                for i in range(len(data_dict)):
                    stage_name = data_dict[i]['system_name']
                    if stage_name:
                        fname = data_dict[i]['form']
                        fid = data_dict[i].pop('form_id')
                        req_action = data_dict[i].pop('required_action_buttons')
                        opt_action = data_dict[i].pop('optional_action_buttons')
                        form_rec = Form.objects.filter(Q(system_name=fname, id=fid)|Q(system_name=fname))
                        if form_rec:
                            form_id = form_rec.values()[0]['id']
                            data_dict[i]['form']=form_id
                            stage_rec = Stage.objects.filter(system_name = stage_name)
                            if stage_rec:
                                sid = stage_rec.values()[0]['id']
                            else:
                                serializer=StageSerializer(data=data_dict[i],context={'request': request})
                                if serializer.is_valid(raise_exception=True):
                                    serializer.save()
                                    count += 1
                                    
                                stage_rec = Stage.objects.get(id=serializer.data.get('id'), system_name = stage_name)
                                sid=stage_rec.id
                            
                            #to save in formstage
                            fstage_rec = FormStage.objects.filter(form = form_id, stage = sid)
                            if not fstage_rec:
                                serializer = FormStage.objects.create(form=form_id, stage = sid)

                            if req_action:
                                text_split = req_action.split(',')
                                for i in range (len(text_split)):
                                    check = StageAction.objects.filter(stage=sid,action=text_split[i])
                                    if not check:
                                        StageAction.objects.create(stage=sid,action=text_split[i],required=True)
                            if opt_action:
                                text_split = opt_action.split(',')
                                for i in range (len(text_split)):
                                    check = StageAction.objects.filter(stage=sid,action=text_split[i])
                                    if not check:
                                        StageAction.objects.create(stage=sid,action=text_split[i],optional=True)
                        else:
                            defective_data.append(fname)
                if defective_data:
                    defective_data = {
                        "missing_form" : f"These {set(defective_data)} are the invalid form names."
                    }
                    return Response(utils.success_def(count,defective_data))
                else:
                    return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))


class StageActionViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Configuration to be modified.
    """
    queryset = StageAction.objects.all()
    serializer_class = StageActionSerializer       
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

 
class ConfigurationViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Configuration to be modified.
    """
    queryset = Configuration.objects.all()
    serializer_class = ConfigurationSerializer       
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    conf_name = data_dict[i]["configuration"]
                    editable = data_dict[i]["editable"]
                    if conf_name:
                        conf = Configuration.objects.filter(system_name=conf_name)
                        if not conf:
                            if editable == 'yes' or 'Yes':
                                data_dict[i]["editable"]=True
                            else:
                                data_dict[i]["editable"] = False
                            serializer=ConfigurationSerializer(data=data_dict[i],context={'request': request})
                            if serializer.is_valid():
                                serializer.save()
                                count += 1
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))
            
class TerritoriesViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Territories to be modified.
    """
    queryset = Territories.objects.all()
    serializer_class = TerritoriesSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

class SelectorViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Selector to be modified.
    """
    queryset = Selectors.objects.all()
    serializer_class = SelectorSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    selector = data_dict[i].get('system_name')
                    if selector:
                        try:
                            serializer=SelectorSerializer(data=data_dict[i], context={'request':request})
                            if serializer.is_valid(raise_exception=True):
                                serializer.save()
                                count += 1
                        except Exception as e:
                            print("Selectors error >>> ", str(e))
                            pass          
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))
            
class ChoiceViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Choice to be modified.
    """
    queryset = Choice.objects.all()
    serializer_class = ChoiceSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    selector = data_dict[i]['selector']
                    choice = data_dict[i]["system_name"]
                    default = data_dict[i]["default"]
                    if selector:
                        selector = utils.encode_api_name(selector)
                        check = Selectors.objects.filter(system_name=selector)
                        if check:
                            data_dict[i]['selector'] = check.values()[0]['id']
                            selector = check.values()[0]['id']
                        else:
                            new_selector = Selectors.objects.create(system_name=selector, type='user')
                            selector = new_selector.id
                            data_dict[i]['selector'] = new_selector.id
                        if choice:
                            choice_name=utils.encode_api_name(choice)
                            choice_rec = Choice.objects.filter(system_name=choice_name, selector = selector)
                            if not choice_rec:
                                if default == 'yes' or 'Yes':
                                    data_dict[i]['default'] = True
                                else:
                                    data_dict[i]['default'] = False
                                try:
                                    serializer = ChoiceSerializer(data=data_dict[i], context={'request':request})
                                    if serializer.is_valid():
                                        serializer.save()
                                        count += 1
                                    # choice_id = Choice.objects.get(id=serializer.data.get('id'))
                                    # language = get_current_user_language(request.user)
                                    # lang = Language.objects.get(name=language)
                                    # check=Translation.objects.filter(label=choice,language_id=lang.id)
                                    # if not check:
                                    #     trans_id = get_rid_pkey('translation')
                                    #     new_label = Translation.objects.create(id=trans_id,label=choice,language_id=lang.id)
                                    # label_rec = Translation.objects.get(label=choice,language_id=lang.id)
                                    # trans = TranslationChoice.objects.filter(choice_id=choice_id.id, translation_id=label_rec.id)
                                    # if not trans:
                                    #     TranslationChoice.objects.create(choice_id=choice_id.id, translation_id=label_rec.id)
                                except Exception as e:
                                    #print(str(e), data_dict[i])
                                    pass
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))

class FormViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Form to be modified.
    """
    queryset = Form.objects.all()
    serializer_class = FormSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    form_name = data_dict[i]['form']
                    if form_name:
                        form_rec = Form.objects.filter(system_name=form_name)
                        if not form_rec:
                            try:
                                serializer=FormSerializer(data=data_dict[i],context={'request': request})
                                if serializer.is_valid(raise_exception=True):
                                    serializer.save()
                                    count += 1
                            except Exception as e:
                                print(str(e))
                                pass
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))

class ListViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows List to be modified.
    """
    queryset = List.objects.all()
    serializer_class = ListSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")
    
    # Related List 
    @action(detail=True, methods=['get'], url_path = "columns")
    def get_columns(self, request, pk=None): 
        queryset = Column.objects.filter(list = pk) 
        serializer = ColumnsSerializer(queryset, many = True, context={'request': request})         
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        try:
            file = request.FILES.get('file')
            selector_id = Selectors.objects.get(selector='list_type')
            if file:
                data_dict = extracting_data(file)
                count = 0
                for i in range(len(data_dict)):
                    try:
                        list_data = data_dict[i]['system_name']
                        category=data_dict[i].pop('category')
                        label = data_dict[i].pop('translation - us english')
                        data_source = data_dict[i].pop('dataset')
                        data_source_id=data_dict[i].pop('dataset_id')
                        list_type = data_dict[i]['list_type']
                        #visibility = data_dict[i]['visibility']
                        lst_id = data_dict[i]['id']
                        table_rec = DataTable.objects.filter(Q(system_name=data_source, id=data_source_id)| Q(system_name=data_source))
                        list_type_rec = Choice.objects.filter(selector=selector_id.id, system_name=list_type)
                        # if visibility:
                        #     visibility_rec = Choice.objects.filter(choice_name=visibility)
                        if list_data and table_rec and list_type_rec:
                            list_rec = List.objects.filter(system_name=list_data)
                            if list_rec:
                                list_id=list_rec.values()[0]['id']
                            else:
                                data_dict[i]['data_source']=table_rec.values()[0]['id']
                                data_dict[i]['list_type']=list_type_rec.values()[0]['id']
                                # if visibility:
                                #     data_dict[i]['visibility']=visibility_rec.values()[0]['id']
                                list_serializers = ListSerializer(data=data_dict[i], context={'request': request})
                                if list_serializers.is_valid(raise_exception=True):
                                    list_serializers.save(id=lst_id)
                                    count = count + 1    
                                list_id=list_serializers.data.get('id')
                            language = get_current_user_language(request.user)
                            lang = Language.objects.get(name=language)
                            if label:
                                language_id = lang.id
                                check=Translation.objects.filter(label=label,language_id=language_id)
                                if not check:
                                    trans_id = get_rid_pkey('translation')
                                    new_label = Translation.objects.create(id=trans_id, label=label,language_id=language_id)
                                label_rec = Translation.objects.get(label=label,language_id=language_id)
                                trans = TranslationList.objects.filter(list_id=list_id, translation_id=label_rec.id)
                                if not trans:
                                    TranslationList.objects.create(list_id=list_id, translation_id=label_rec.id)
                            if category:
                                menu_rec = Menu.objects.filter(menu_category=category,list_id=list_id)
                                if not menu_rec:
                                    mdict = {}
                                    mdict['menu_category']=category
                                    mdict['list']=list_id
                                    mrec=MenuSerializer(data=mdict, context={'request': request})
                                    if mrec.is_valid():
                                        mrec.save()
                    except Exception as e:
                        print(str(e))
                        pass
                return Response(utils.success(count))
            else:
                msg="Please Upload A Suitable Excel File."
                return Response(utils.error(msg))
        except Exception as e:
            return Response(utils.error(str(e)))

class ListFiltersViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows ListFilters to be modified.
    """
    queryset = ListFilters.objects.all()
    serializer_class = ListFilterSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

class ColumnsViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Columns to be modified.
    """
    queryset = Column.objects.all()
    serializer_class = ColumnsSerializer  
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    # filterset_fields = ("app__name",)
    filterset_fields = {
            'position': ['exact'],'visibility': ['exact']
        }
    ordering_fields = ("__all__")
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        file = request.FILES.get('file')
        if file:
            data_dict = extracting_data(file)
            count = 0
            defective_data=[]
            for i in range(len(data_dict)):
                column = data_dict[i]['system_name']
                field = data_dict[i]['field']
                list = data_dict[i]['list']
                visibility = data_dict[i]['visibility']
                list_rec = List.objects.filter(system_name=list)
                visibility_rec = Choice.objects.filter(system_name=visibility)
                try:
                    if column and field and list_rec and visibility_rec:
                        list_id = list_rec.values()[0]['id']
                        data_dict[i]['list'] = list_id
                        columnrec = Column.objects.filter(system_name = column,list_id=list_id)
                        if not columnrec:
                            serializer = ColumnsSerializer(data=data_dict[i],context={'request': request})
                            if serializer.is_valid(raise_exception=True):
                                serializer.save()
                                count = count+1
                except Exception as e:
                    print("Column Error >> ", str(e))
                    pass
                if not list_rec:
                    defective_data.append(list)
            if defective_data:
                defective_data = {
                    "missing_lists" : f"These {defective_data} are the invalid list names."
                } 
                return Response(utils.success_def(count,defective_data))
            else:
                return Response(utils.success(count))
        else:
            msg="Please Upload A Suitable Excel File."
            return Response(utils.error(msg))
        
class MenuViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Country to be modified.
    """
    queryset = Menu.objects.all()
    serializer_class = MenuSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    # filterset_fields = ("__all__")
    filterset_fields = {
        'system_name': ['exact', 'contains'],
        'list': ['exact'],
        'sequence': ['exact']
    }
    ordering_fields = ("__all__")
    
    # def list(self, request, *args, **kwargs):
    #     queryset = Menu.objects.filter().all()
    #     serializers = MenuSerializer(queryset, many = True, context = {"request": request})
    #     return super().list(request, *args, **kwargs)
    
    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        file = request.FILES.get('file')
        if file:
            data_dict = extracting_data(file)
            count = 0
            defective_data = []
            for i in range(len(data_dict)):
                menu_name = data_dict[i]['system_name']
                if menu_name:
                    list = data_dict[i]['list']
                    lst_id = data_dict[i].pop('list_id')
                    category=data_dict[i]['menu_category']
                    category_id=data_dict[i].pop('menu_category_id')
                    if list:
                        listrec = List.objects.filter(Q(system_name__contains=list, id=lst_id)|Q(id=lst_id))
                        if listrec:
                            data_dict[i]['list'] = listrec.values()[0]['id']
                            list_id = listrec.values()[0]['id']
                            menu_rec = Menu.objects.filter(system_name__contains=menu_name, list_id = list_id)
                        else:
                            defective_data.append(list)   
                    else:
                        data_dict[i]['list'] = None
                        menu_rec = Menu.objects.filter(system_name__contains=menu_name)
                    if category:
                        choice_id=Choice.objects.filter(system_name__contains=category, id=category_id)
                        if choice_id:
                            data_dict[i]['menu_category']=choice_id.values()[0]['id']
                    language = get_current_user_language(request.user)
                    lang = Language.objects.get(system_name__contains=language)
                    check=Translation.objects.filter(label__contains=menu_name, language_id=lang.id)
                    if not check:
                        trans_id = get_rid_pkey('translation')
                        new_label = Translation.objects.create(id=trans_id, label=menu_name, language_id=lang.id)
                    label_rec = Translation.objects.get(label=menu_name, language_id=lang.id)
                    if not menu_rec:
                        try:
                            menu_serializer = MenuSerializer(data=data_dict[i], context={'request': request})
                            if menu_serializer.is_valid(raise_exception=True):
                                menu_serializer.save()
                                count = count + 1
                                new_trans = TranslationMenu.objects.create(menu_id=menu_serializer.data.get('id'), translation_id=label_rec.id)
                        except Exception as e:
                            print("menu Error >>> ", str(e))
            if defective_data:
                defective_data = {
                    "invalid_lists" : f"These {set(defective_data)} are the invalid list name(s)."
                } 
                return Response(utils.success_def(count,defective_data))
            else:
                return Response(utils.success(count))
        else:
            msg="Please Upload A Suitable Excel File."
            return Response(utils.error(msg))  
     
class HelpViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Country to be modified.
    """
    queryset = Help.objects.all()
    serializer_class = HelpSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

class FormListViewSet(viewsets.ModelViewSet):
    queryset = FormList.objects.all()
    serializer_class = FormListSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        file = request.FILES.get('file')
        if file:
            data_dict = extracting_data(file)
            count = 0
            for i in range(len(data_dict)):
                list=data_dict[i]['list']
                form = data_dict[i]['form']
                if list and form:
                    try:
                        formrec = Form.objects.get(system_name=form)
                        listrec = List.objects.get(system_name__contains=list)
                        if formrec and listrec:
                            formlistrec = FormList.objects.filter(list=listrec.id,form=formrec.id)
                            data_dict[i]['form']= formrec.id
                            data_dict[i]['list']= listrec.id
                            primary = data_dict[i]['primary']
                            if primary == 'yes' or 'Yes':
                                data_dict[i]['primary'] = True
                            else:
                                data_dict[i]['primary'] = False

                            if not formlistrec:
                                serializer=FormListSerializer(data=data_dict[i], context={'request':request})
                                if serializer.is_valid(raise_exception=True):
                                    serializer.save()
                                    count += 1
                    except Exception as e:
                        print(str(e), data_dict[i])
                        pass
            return Response(utils.success(count))
        else:
            msg="Please Upload A Suitable Excel File."
            return Response(utils.error(msg))

class FormDataViewSet(viewsets.ModelViewSet):
    queryset = FormData.objects.all()
    serializer_class = FormDataSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

    @action(detail=False, methods=['post'], name='import_data', url_path = "import")
    def import_data(self, request):
        file = request.FILES.get('file')
        if file:
            data_dict = extracting_data(file)
            count = 0
            for i in range(len(data_dict)):
                data=data_dict[i]['data']
                data=Data.objects.filter(system_name=data)
                field = data_dict[i]['field']
                table = data_dict[i]['table']
                table = DataTable.objects.filter(system_name= table)
                form = data_dict[i]['form']
                form = Form.objects.filter(system_name=form)
                try:
                    if data and field and table and form:
                        table_id = table.values()[0]['id']
                        form_id = form.values()[0]['id']
                        data_id = data.values()[0]['id']
                        fd_rec = FormData.objects.filter(table=table_id, data=data_id, form=form_id)
                        if not fd_rec:
                            data_dict[i]['form']=form_id
                            data_dict[i]['data']=data_id
                            data_dict[i]['table']=table_id
                            data_dict[i]['type']=(data_dict[i]['type']).lower()
                            serializer=FormDataSerializer(data=data_dict[i], context={'request':request})
                            if serializer.is_valid(raise_exception=True):
                                serializer.save()
                                count += 1
                except Exception as e:
                    print("Form Data Error >>> ", str(e))
                    pass
            return Response(utils.success(count))
        else:
            msg="Please Upload A Suitable Excel File."
            return Response(utils.error(msg))

class FormSectionViewSet(viewsets.ModelViewSet):
    queryset = FormSection.objects.all()
    serializer_class = RelatedFormSectionSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

class IconViewSet(viewsets.ModelViewSet):
    queryset = Icons.objects.all()
    serializer_class = IconSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("system_name",)
    ordering_fields = ("system_name",)
    
class ActionViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Actions to be modified.
    """
    queryset = Action.objects.all()
    serializer_class = ActionSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

class FormStageViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows FormStage to be modified.
    """
    queryset = FormStage.objects.all()
    serializer_class = FormStageSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")

class ButtonStageViewSet(viewsets.ModelViewSet):
    """
    API’s endpoint that allows Button Stage to be modified.
    """
    queryset = ButtonStage.objects.all()
    serializer_class = ButtonStageSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ("__all__")
    ordering_fields = ("__all__")